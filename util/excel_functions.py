import openpyxl
import datetime


class ExcelFunctions:

    def get_number_rows(self, file: str, sheet: str):
        wordbook = openpyxl.load_workbook(file)
        sheet = wordbook[sheet]
        return sheet.max_row

    def get_number_columns(self, file: str, sheet: str):
        wordbook = openpyxl.load_workbook(file)
        sheet = wordbook[sheet]
        return sheet.max_row

    def read_data(self, file, sheet, row_num: int, column_num: int):
        """
        Allow to get data of a cell.
        :return: str
        """
        wordbook = openpyxl.load_workbook(file)
        sheet = wordbook[sheet]
        return sheet.cell(row=row_num, column=column_num).value

    def write_data(self, file: str, sheet: str, row_num: int, column_num: int, data):
        wordbook = openpyxl.load_workbook(file)
        sheet = wordbook[sheet]
        sheet.cell(row=row_num, column=column_num).value = data
        wordbook.save(file)

    def get_data(self, file: str, sheet: str):
        """
        Allow to get data from an Excel file.
        :return: list[list[]]
        """
        try:
            cell_ref_i = ""
            column_ref_f = ""
            validate_i = False
            validate_f = False
            data_list = []
            excel_file = openpyxl.load_workbook(file)
            # Sheet name
            sheet_name = excel_file[sheet]
            # Number of rows and columns
            max_rows = sheet_name.max_row
            max_columns = sheet_name.max_column
            # Going through Excel until you find the references
            for i in range(max_rows):
                for j in range(max_columns):
                    cell = sheet_name.cell(i + 1, j + 1)
                    if cell.value == "N":
                        cell_ref_i = cell.offset(row=1, column=0)
                        column_ref_i = cell_ref_i.column_letter
                        validate_i = True
                    if cell.value == "RESULT":
                        cell_ref_f = cell.offset(row=1, column=-1)
                        column_ref_f = cell_ref_f.column_letter
                        validate_f = True
            if not validate_i:
                print("Could not find field 'N'")
                return False, []
            if not validate_f:
                print("Could not find field 'Result'")
                return False, []
            data = sheet_name[cell_ref_i.coordinate:column_ref_f + str(max_rows)]
            for row in data:
                row_data = []
                cell_number = row[0].value
                if isinstance(cell_number, int):
                    for cell in row[0:len(row)]:
                        field_value = cell.value
                        if field_value is None:
                            row_data.append("")
                        else:
                            if isinstance(field_value, datetime.datetime):
                                field_value = field_value.strftime("%d/%m/%Y")
                            row_data.append(field_value)
                if row_data:
                    data_list.append(row_data)
            return data_list
        except Exception as e:
            print(e)
            print("Error getting data from Excel file.")

    def write_data_to_a_cell(self, file: str, sheet: str, row_num: int, data: str):
        """
        Allow writing text in a cell of an Excel file.
        :return: bool
        """
        try:
            cell_ref = ""
            validate = False
            excel_file = openpyxl.load_workbook(file)
            # Sheet name
            sheet_name = excel_file[sheet]
            # Number of rows and columns
            max_rows = sheet_name.max_row
            max_columns = sheet_name.max_column
            # Going through Excel until you find the references
            for i in range(max_rows):
                for j in range(max_columns):
                    cell = sheet_name.cell(i + 1, j + 1)
                    if cell.value == "RESULT":
                        cell_ref = cell.offset(row=row_num)
                        validate = True
            if not validate:
                print("Could not find field 'Result'")
                return False
            cell_ref.value = data
            excel_file.save(file)
            return True
        except Exception as e:
            print(e)
            print("Error saving data to Excel file.")